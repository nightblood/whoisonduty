# -*- coding: utf-8 -*-

import sys

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from functools import reduce
import datetime
from openpyxl import Workbook
import random
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.comments import Comment
import json
from urllib import request
import math
import qdarkstyle
from logzero import logger

from mask_layout import MaskWidget
from staff import StaffList


class MainWindow(QWidget):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        MainWindow.setFixedSize(self, 760, 500)
        self.setWindowTitle('排班工具')
        self.setGeometry(500, 300, 300, 200)

        self.export_worker = ''
        self.start_dt = ''
        self.end_dt = ''
        self.staff = []
        self.init_data()
        self.init_window()

    def init_data(self):
        self.init_staff()
        self.init_date()
        self.next_holiday = self.get_next_holiday()

    def get_next_holiday(self):
        '''
        response：{"code":0,"tts":"最近的一个节日是2022-01-01的元旦，还有61天。"}
        :return:
        '''
        try:
            url = "http://timor.tech/api/holiday/tts/next"
            resp = request.urlopen(url, data=None, timeout=5)
            data = json.loads(resp.read())
            if data["code"] == 0:
                print(data["tts"])
                return data["tts"]
        except Exception as e:
            logger.error(e)
            return ""

    def init_window(self):
        self.main_layout = QVBoxLayout(self)
        self.l_desc = QLabel('将所有名字填入下方输入框，并以空格隔开。点击【排班】按钮将随机排列名字。')
        self.pte_staff = QPlainTextEdit(self.get_staff_str())
        self.pte_staff.setFont(QFont('Roman times', 14, QFont.Bold))
        self.lb_holiday_desc = QLabel(self.next_holiday)
        self.l_start_dt = QLabel('开始日期')
        self.l_end_dt = QLabel('结束日期')

        self.dte_start = QDateTimeEdit(QDate.currentDate(), self)
        self.dte_start.dateTimeChanged.connect(lambda: self.set_time(1))
        self.dte_start.setDisplayFormat('yyyy年MM月dd日')
        self.dte_start.setCalendarPopup(True)
        self.dte_end = QDateTimeEdit(QDate.currentDate().addDays(7 * 40), self)
        self.dte_end.dateTimeChanged.connect(lambda: self.set_time(2))
        self.dte_end.setDisplayFormat('yyyy年MM月dd日')
        self.dte_end.setCalendarPopup(True)

        self.layout_start_dt = QSplitter(Qt.Horizontal)
        self.layout_end_dt = QSplitter(Qt.Horizontal)

        self.layout_func = QSplitter(Qt.Horizontal)
        self.btn_sort = QPushButton('排班')
        self.btn_export = QPushButton('导出')
        self.layout_func.addWidget(self.btn_sort)
        self.layout_func.addWidget(self.btn_export)

        self.l_weeks = QLabel('共 40 周')

        self.layout_start_dt.addWidget(self.l_start_dt)
        self.layout_start_dt.addWidget(self.dte_start)
        self.layout_end_dt.addWidget(self.l_end_dt)
        self.layout_end_dt.addWidget(self.dte_end)

        self.btn_export.clicked.connect(self.on_export)
        self.btn_sort.clicked.connect(self.on_sort)

        self.layout_miss = QSplitter(Qt.Horizontal)
        self.r_btn_miss = QRadioButton("是否需要轮空计算")
        self.layout_miss.addWidget(self.r_btn_miss)

        self.main_layout.addWidget(self.l_desc)
        self.main_layout.addWidget(self.pte_staff)
        self.main_layout.addWidget(self.layout_miss)
        self.main_layout.addWidget(self.layout_start_dt)
        self.main_layout.addWidget(self.layout_end_dt)
        self.main_layout.addWidget(self.l_weeks)
        self.main_layout.addWidget(self.lb_holiday_desc)
        self.main_layout.addWidget(self.layout_func)

    def set_time(self, btn_idx):
        if btn_idx == 1:
            self.start_dt = self.dte_start.text()
        else:
            self.end_dt = self.dte_end.text()
        print(self.start_dt, self.end_dt)
        self.l_weeks.setText('共' + str(self.get_delta_weeks()) + '周')

    def on_export(self):
        self.staff = self.get_staff_list_from_widget()
        weeks = self.get_delta_weeks()
        if self.staff is None or len(self.staff) == 0:
            QMessageBox.question(self, '提示', '请先输入人员，再进行排班', QMessageBox.Yes)
            return
        if weeks <= 0:
            QMessageBox.question(self, '提示', '日期选择异常，请重新选择日期', QMessageBox.Yes)
            return
        self.export_worker = ExportWorker(self.staff, weeks, self.start_dt, self.r_btn_miss.isChecked())
        self.export_worker.started.connect(self.show_mask)
        self.export_worker.sig_complete.connect(self.export_complete)
        self.export_worker.finished.connect(self.hide_mask)
        self.export_worker.start()

    def show_mask(self, msg="初始化中。。。"):
        """
        显示遮罩层，遮罩层可以显示信息msg
        :param msg:
        :return:
        """
        self.mask = MaskWidget(self)
        self.mask.show()
        self.mask.set_msg("制作中。。。")

    def hide_mask(self):
        if self.mask is not None and self.mask != "":
            self.mask.close()

    def export_complete(self, desc):
        try:
            QMessageBox.question(self, '提示', desc, QMessageBox.Yes)
        except Exception as e:
            logger.error(e)

    def get_delta_weeks(self):
        start = datetime.datetime.strptime(self.start_dt, '%Y年%m月%d日')
        end = datetime.datetime.strptime(self.end_dt, '%Y年%m月%d日')
        weeks = (end - start).days / 7
        if (end - start).days % 7 > 0:
            weeks += 1
        return weeks

    def on_sort(self):
        self.staff = self.get_staff_list_from_widget()
        if self.staff is None or len(self.staff) == 0:
            QMessageBox.question(self, '提示', '请先输入人员，再进行排班', QMessageBox.Yes)
            return
        self.sort_worker = SortWorker(self.staff)
        self.sort_worker.sig_complete.connect(self.sort_complete)
        self.sort_worker.start()

    def sort_complete(self, data):
        self.staff = data.split(' ')
        self.pte_staff.setPlainText(self.get_staff_str())

    def init_staff(self):
        try:
            staff_file = open('staff.txt', 'r', encoding='utf-8')
            data = staff_file.read()
            print(data)
            if data is not None and len(data.strip()) != 0:
                for item in data.strip().split(' '):
                    if len(item) != 0:
                        self.staff.append(item)
            staff_file.close()
        except Exception as e:
            logger.error(e)

    def init_date(self):
        now = datetime.datetime.now()
        self.start_dt = now.strftime('%Y{y}%m{m}%d{d}').format(y='年', m='月', d='日')
        # 40周
        self.end_dt = (now + datetime.timedelta(days=280)).strftime('%Y{y}%m{m}%d{d}').format(y='年', m='月', d='日')

    def get_staff_str(self):
        if self.staff is None or len(self.staff) == 0:
            return ''
        return reduce(lambda x, y: x + ' ' + y, self.staff)

    def get_staff_list_from_widget(self):
        res = []
        str = self.pte_staff.toPlainText()
        if str is None or len(str.strip()) == 0:
            return res
        staffs = str.strip().split(' ')
        print(staffs)
        for item in staffs:
            if len(item) != 0:
                res.append(item)
        return res


class SortWorker(QThread):
    sig_complete = pyqtSignal(str)

    def __init__(self, staff):
        super().__init__()
        self.staff = staff

    def run(self):
        length = len(self.staff)
        res = []
        for idx in range(length):
            item = random.choice(self.staff)
            self.staff.remove(item)
            res.append(item)
        print(res)
        self.sig_complete.emit(reduce(lambda x, y: x + ' ' + y, res))


class ExportWorker(QThread):
    sig_complete = pyqtSignal(str)

    def __init__(self, staffs, weeks, start_dt, miss_flag):
        super().__init__()

        self.staff = StaffList(staffs, miss_flag)
        self.weeks = weeks
        self.start_dt = start_dt

    def run(self):
        try:
            wb = Workbook()
            ws = wb.active
            border = Border(left=Side(border_style='thin', color='4f5555'), right=Side(border_style='thin', color='4f5555'),
                            top=Side(border_style='thin', color='4f5555'), bottom=Side(border_style='thin', color='4f5555'))
            pattern_fill = PatternFill("solid", fgColor="feeeed")  # 星期行和日期列背景色
            bg_weekend = PatternFill("solid", fgColor="fedcbd")  # 周末值班人背景色
            bg_holidayd = PatternFill("solid", fgColor="f15b6c")  # 节假日
            bg_holidayd_work = PatternFill("solid", fgColor="87843b")  # 节假日调休
            col_name = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

            self.init_excel(ws)  # 表头

            # 第一行为【值班表】第二行为【星期】， 从第三行开始排班
            rol_start_idx = 3
            # 计算总行数
            row_number = math.ceil(self.weeks) + rol_start_idx - 1
            # 找到周一的日期，和对应的人
            start_time = datetime.datetime.strptime(self.start_dt, '%Y年%m月%d日')
            delta = start_time.weekday()
            start_time = start_time - datetime.timedelta(days=delta)
            # 找到周一值班的人，
            # staff_idx = self.staff.cnt_staff - delta

            holidays = self.get_holidays(start_time, self.weeks)

            for row_idx in range(rol_start_idx, int(row_number) + rol_start_idx - 1):  # 从第三行开始排班 [3, 4, 5...]
                ws.row_dimensions[row_idx].height = 30  # 行高
                # 每个单元格批注
                date_desc = (start_time + datetime.timedelta(weeks=row_idx - rol_start_idx)).strftime('%Y{y}%m{m}%d{d}')\
                    .format(y='年', m='月', d='日') + "-" \
                    + (start_time + datetime.timedelta(weeks=row_idx - rol_start_idx, days=6))\
                    .strftime('%m{m}%d{d}').format(m='月', d='日')

                for col_idx in range(len(col_name)):
                    cell_pos = col_name[col_idx] + str(row_idx)
                    bg_cell = None
                    if col_idx == 0:
                        ws[cell_pos] = date_desc
                        ws[cell_pos].fill = pattern_fill
                    else:
                        if col_idx <= delta and row_idx == rol_start_idx:
                            ws[cell_pos] = ""  # 第三行第一个人前面表格为空（不排班）只排指定日期开始之后的
                            logger.debug("还没开始。。。")
                        else:
                            staff = self.staff.get_staff_avilable()
                            ws[cell_pos] = staff

                        date_item = (start_time + datetime.timedelta(weeks=row_idx - 3, days=col_idx - 1)).strftime(
                            '%Y{y}%m{m}%d{d}').format(y='-', m='-', d='')
                        comment = date_item
                        if col_idx in (6, 7):
                            bg_cell = bg_weekend

                        # 添加节假日批注
                        if holidays is not None and holidays.__contains__(date_item) \
                                and holidays[date_item] is not None \
                                and holidays[date_item]["holiday"] is not None:
                            if holidays[date_item]["holiday"]:
                                bg_cell = bg_holidayd
                            else:
                                bg_cell = bg_holidayd_work
                            comment = holidays[date_item]["name"] + "\n" + holidays[date_item]["date"]

                        ws[cell_pos].comment = Comment(comment, 'zlf')
                    if bg_cell is not None:
                        ws[cell_pos].fill = bg_cell
                    ws[cell_pos].alignment = Alignment(horizontal='center', vertical='center')
                    ws[cell_pos].border = border
            file = '值班表' + self.start_dt + '.xlsx'
            wb.save(file)
            self.update_staff_file()
            self.sig_complete.emit('已在当前目录下导出值班表 ' + file)
        except Exception as e:
            logger.error(e)
            self.sig_complete.emit('异常：' + str(e))

    def init_excel(self, ws):
        ws.row_dimensions[1].height = 40  # 第一行行高
        ws.row_dimensions[2].height = 30  # 第二行行高
        ws.merge_cells('A1:H1')  # 合并单元格
        ws['A1'] = "值 班 表"
        ws["A1"].fill = PatternFill("solid", fgColor="76becc")
        ws["A1"].border = Border(left=Side(border_style='thin', color='4f5555'), right=Side(border_style='thin', color='4f5555'),
                                 top=Side(border_style='thin', color='4f5555'), bottom=Side(border_style='thin', color='4f5555'))
        ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
        ws["A1"].font = Font(u'宋体', size=20, bold=True, italic=False, strike=False, color='000000')

        col_name = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        first_row = ['日期', '周一', '周二', '周三', '周四', '周五', '周六', '周日']


        # 初始化星期行
        for col_idx in range(len(col_name)):
            col = col_name[col_idx]
            cell_pos = col + "2"
            ws[cell_pos] = first_row[col_idx]
            ws[cell_pos].alignment = Alignment(horizontal='center', vertical='center')
            ws[cell_pos].fill = PatternFill("solid", fgColor="feeeed")
            ws[cell_pos].border = Border(left=Side(border_style='thin', color='4f5555'), right=Side(border_style='thin', color='4f5555'),
                                         top=Side(border_style='thin', color='4f5555'), bottom=Side(border_style='thin', color='4f5555'))
            if col_idx == 0:
                ws.column_dimensions[col].width = 25
            else:
                ws.column_dimensions[col].width = 15

    def get_holidays(self, start_time, weeks):
        try:
            items = start_time.strftime('%Y-%m-%d')

            for i in range(int(weeks * 7) + 1):
                items = items + "," + (start_time + datetime.timedelta(days=i)).strftime('%Y{y}%m{m}%d{d}').format(y='-', m='-', d='')
            print(items)

            url = "http://timor.tech/api/holiday/batch?d=" + items
            resp = request.urlopen(url, data=None, timeout=5)
            data = json.loads(resp.read())
            if data["code"] == 0:
                print(data["holiday"])
                return data["holiday"]
        except Exception as e:
            logger.error(e)

    def update_staff_file(self):
        try:
            staff_file = open('staff.txt', 'w', encoding='utf-8')
            staff_file.write(reduce(lambda x, y: x + ' ' + y, self.staff.names))
            staff_file.close()
        except Exception as e:
            logger.error(e)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()

    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())

    icon = QIcon()
    icon.addPixmap(QPixmap('icon.ico'), QIcon.Normal, QIcon.Off)
    window.setWindowIcon(icon)

    window.show()
    sys.exit(app.exec_())
